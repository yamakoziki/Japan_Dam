"""
全国ダム地質DB 地質推定可能性分析 v2
=====================================
symbol記号に基づく複層地質推定可能性を分析し、データ充実が望ましいダムを提案する。

A1_北海道地質推定分析     : 北海道ダムの複層地質推定可能性（層別カバレッジ・分布）
A2_全国地質推定分析       : 北海道以外ダムの複層地質推定可能性（都道府県別内訳付き）
A3_情報充実推奨100ダム    : symbol空白・不足ダムの優先充実提案リスト（100件）
A4_管理者別カバレッジ     : 管理者機関ごとの推定可能性（充実優先機関の特定）
A5_都道府県×地質層マトリクス : 都道府県×Layer カバレッジ熱地図（地域×地質時代の空白可視化）
A6_層別空白補完分析       : 各地質層ごとの空白補完優先ダム（層単位の充実ターゲット）

北海道ダム定義 : 所在地が「北海道」で始まるすべてのダム
全国ダム定義   : 所在地が「北海道」以外のダム
推定可能性評価 : 0層=推定不可  1〜2層=推定限定的  3〜5層=推定良好
"""

import argparse
import datetime
import sys
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── スタイル ─────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUB_FILL  = PatternFill("solid", fgColor="2E75B6")
SUB_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SEC_FILL  = PatternFill("solid", fgColor="D6E4F0")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
ALT_FILL  = PatternFill("solid", fgColor="F2F7FC")
RED_FILL  = PatternFill("solid", fgColor="FFE0E0")
YEL_FILL  = PatternFill("solid", fgColor="FFF8DC")
GRN_FILL  = PatternFill("solid", fgColor="E8F5E9")
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right",  vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── 定数 ─────────────────────────────────────────────────────
LAYER_ID_COL = {1: 23, 2: 35, 3: 47, 4: 59, 5: 71}
LAYER_NAMES = {
    1: "Layer1: Pre-N（新第三紀以前）",
    2: "Layer2: N-古（新第三紀・古い方）",
    3: "Layer3: N-新（新第三紀・新しい方）",
    4: "Layer4: Q-old（中期更新世）",
    5: "Layer5: Q-H（後期更新世〜完新世）",
}
LAYER_SHORT = {1: "Pre-N", 2: "N-古", 3: "N-新", 4: "Q-old", 5: "Q-H"}

BEARING_SCORE = {
    "低": 1, "低〜中": 2, "中": 3,
    "中〜高": 4, "中〜高（溶結度依存）": 4,
    "高（続成固結）": 5, "高": 5,
}
PERM_SCORE = {
    "低〜中": 1, "中（節理閉鎖）": 2, "中": 2, "中〜高": 3,
    "高": 4, "高（柱状節理）": 4, "高（柱状節理・カルデラ）": 4,
}

def b_score(v): return BEARING_SCORE.get(v, 0)
def p_score(v): return PERM_SCORE.get(v, 0)


# ─── スタイルヘルパー ─────────────────────────────────────────
def hdr(ws, r, c, v, w=None):
    cell = ws.cell(r, c, v)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = CTR; cell.border = BDR
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cell

def sub(ws, r, c, v, w=None):
    cell = ws.cell(r, c, v)
    cell.font = SUB_FONT; cell.fill = SUB_FILL
    cell.alignment = CTR; cell.border = BDR
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cell

def body(ws, r, c, v, align=LEFT, bold=False, fill=None):
    cell = ws.cell(r, c, v)
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.alignment = align; cell.border = BDR
    if fill: cell.fill = fill
    return cell

def sec_title(ws, r, c1, c2, text):
    ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
    cell = ws.cell(r, c1, text)
    cell.font = BOLD_FONT; cell.fill = SEC_FILL; cell.alignment = LEFT
    return cell

def sheet_title(ws, r, c1, c2, text):
    ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
    cell = ws.cell(r, c1, text)
    cell.font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    cell.alignment = LEFT
    return cell

def pct(n, total): return round(n / total * 100, 1) if total else 0
def avg(lst): return round(mean(lst), 2) if lst else ""
def scol(c): return get_column_letter(c)


# ─── データ読み込み ───────────────────────────────────────────
def load_data(wb):
    ws_db = wb["全国ダム地質DB"]
    ws_gl = wb["Glossary"]

    glossary = {}
    fields = ["id", "symbol", "geo_surface", "geo_era", "geo_rock",
              "formationAge_ja", "group_ja", "lithology_ja", "geo_rock_label",
              "bearing_cap", "permeability", "main_risk"]
    for row in ws_gl.iter_rows(min_row=4, values_only=True):
        if row[0] is None: continue
        try:
            rec = {f: row[i] for i, f in enumerate(fields)}
            glossary[int(rec["id"])] = rec
        except:
            continue

    dams = []
    for row in range(3, ws_db.max_row + 1):
        name = ws_db.cell(row, 3).value
        if not name: continue

        layers = {}
        for lnum, col in LAYER_ID_COL.items():
            gid = ws_db.cell(row, col).value
            if gid is not None and not isinstance(gid, str):
                try:
                    layers[lnum] = glossary.get(int(gid))
                except:
                    layers[lnum] = None
            else:
                layers[lnum] = None

        filled = [k for k, v in layers.items() if v is not None]
        recs = [v for v in layers.values() if v is not None]

        loc = ws_db.cell(row, 17).value or ""
        mgr = ws_db.cell(row, 14).value
        height = ws_db.cell(row, 10).value

        dams.append({
            "name": name,
            "loc": loc,
            "pref": loc[:3] if loc else "",
            "mgr_code": mgr,
            "height": height if isinstance(height, (int, float)) else None,
            "layers": layers,
            "recs": recs,
            "filled_layers": filled,
            "layer_count": len(filled),
        })
    return dams, glossary


# ─── 推定可能性 ───────────────────────────────────────────────
def estimability_label(n):
    if n == 0: return "推定不可"
    if n <= 2: return "推定限定的"
    return "推定良好"

def estimability_fill(n):
    if n == 0: return RED_FILL
    if n <= 2: return YEL_FILL
    return GRN_FILL


# ─── 共通: カバレッジ3セクション（サマリー・層別・層数分布） ──
def _write_coverage_sections(ws, row, target_dams):
    total = len(target_dams)
    with_data = [d for d in target_dams if d["layer_count"] > 0]
    no_data   = [d for d in target_dams if d["layer_count"] == 0]
    good      = [d for d in target_dams if d["layer_count"] >= 3]
    partial   = [d for d in target_dams if 1 <= d["layer_count"] <= 2]

    # ── セクション 1: サマリー ──
    sec_title(ws, row, 1, 3, "▼ カバレッジサマリー")
    row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    for label, val, pct_v in [
        ("総ダム数",                         total,                                   ""),
        ("データあり（1層以上）",              len(with_data),                           f"{pct(len(with_data), total)}%"),
        ("データなし（0層）",                 len(no_data),                             f"{pct(len(no_data), total)}%"),
        ("推定良好（3層以上）",               len(good),                                f"{pct(len(good), total)}%"),
        ("推定限定的（1〜2層）",              len(partial),                             f"{pct(len(partial), total)}%"),
        ("推定不可（0層）",                   len(no_data),                             f"{pct(len(no_data), total)}%"),
        ("平均層数（全ダム）",               avg([d["layer_count"] for d in target_dams]), ""),
        ("平均層数（データありダムのみ）",    avg([d["layer_count"] for d in with_data]) if with_data else "", ""),
    ]:
        ws.cell(row, 1, label).font = BOLD_FONT
        ws.cell(row, 1).alignment  = LEFT
        body(ws, row, 2, val,   align=RIGHT)
        body(ws, row, 3, pct_v, align=RIGHT)
        row += 1
    row += 1

    # ── セクション 2: 層別カバレッジ ──
    sec_title(ws, row, 1, 7, "▼ 層別カバレッジ（各層にデータを持つダム数）")
    row += 1
    for ci, (t, w) in enumerate(zip(
        ["層番号", "層名", "ダム数", "カバー率(%)", "代表Symbol-1", "代表Symbol-2", "代表Symbol-3"],
        [8, 30, 8, 12, 18, 18, 18]
    ), 1):
        hdr(ws, row, ci, t, w)
    row += 1

    for lnum in range(1, 6):
        dams_l = [d for d in target_dams if d["layers"][lnum] is not None]
        top3 = [s for s, _ in Counter(
            d["layers"][lnum]["symbol"]
            for d in dams_l
            if d["layers"][lnum] and d["layers"][lnum].get("symbol")
        ).most_common(3)]
        cr = pct(len(dams_l), total)
        fill = GRN_FILL if cr >= 40 else YEL_FILL if cr >= 10 else RED_FILL
        vals = [lnum, LAYER_NAMES[lnum], len(dams_l), cr] + top3 + [""] * (3 - len(top3))
        for ci, v in enumerate(vals, 1):
            body(ws, row, ci, v, align=RIGHT if ci in (1, 3, 4) else LEFT, fill=fill)
        row += 1
    row += 1

    # ── セクション 3: 層数分布 ──
    sec_title(ws, row, 1, 5, "▼ 層数別ダム分布（複層推定可能性の全体像）")
    row += 1
    for ci, (t, w) in enumerate(zip(
        ["層数", "件数", "割合(%)", "推定可能性評価", "代表所在地（上位3）"],
        [8, 8, 10, 16, 40]
    ), 1):
        hdr(ws, row, ci, t, w)
    row += 1

    for lcount in range(6):
        dams_lc = [d for d in target_dams if d["layer_count"] == lcount]
        if not dams_lc: continue
        pref_top3 = "、".join(p for p, _ in Counter(d["pref"] for d in dams_lc).most_common(3))
        fill = estimability_fill(lcount)
        for ci, v in enumerate(
            [lcount, len(dams_lc), pct(len(dams_lc), total), estimability_label(lcount), pref_top3], 1
        ):
            body(ws, row, ci, v, align=RIGHT if ci in (1, 2, 3) else LEFT, fill=fill)
        row += 1
    row += 1

    return row


# ─── 共通: ダム別詳細テーブル ─────────────────────────────────
def _write_dam_detail(ws, row, target_dams, section_label=None):
    total = len(target_dams)
    label = section_label or f"▼ ダム別詳細一覧（層数降順・堤高降順、全{total}件）"
    sec_title(ws, row, 1, 13, label)
    row += 1

    hdrs   = ["ダム名", "所在地", "堤高(m)", "管理者", "層数",
              "L1-symbol", "L2-symbol", "L3-symbol", "L4-symbol", "L5-symbol",
              "推定可能性", "平均強度", "平均透水性"]
    widths = [16, 20, 8, 8, 6, 16, 16, 16, 16, 16, 14, 10, 10]
    hdr_row = row
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        hdr(ws, row, ci, h, w)
    row += 1

    for d in sorted(target_dams, key=lambda d: (-d["layer_count"], -(d["height"] or 0))):
        fill = estimability_fill(d["layer_count"])
        b_avg = avg([b_score(r["bearing_cap"]) for r in d["recs"] if b_score(r["bearing_cap"])])
        p_avg = avg([p_score(r["permeability"]) for r in d["recs"] if p_score(r["permeability"])])
        layer_syms = [
            d["layers"][n]["symbol"] if d["layers"][n] and d["layers"][n].get("symbol") else ""
            for n in range(1, 6)
        ]
        vals = [d["name"], d["loc"][:20] if d["loc"] else "", d["height"],
                d["mgr_code"], d["layer_count"]] + layer_syms + [
                estimability_label(d["layer_count"]), b_avg, p_avg]
        for ci, v in enumerate(vals, 1):
            body(ws, row, ci, v, align=RIGHT if ci in (3, 4, 5, 12, 13) else LEFT, fill=fill)
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A{hdr_row}:{scol(len(hdrs))}{hdr_row}"
    return row


# ─── A1: 北海道地質推定分析 ─────────────────────────────────
def write_a1(wb, dams, glossary):
    ws = wb.create_sheet("A1_北海道地質推定分析")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 1, 13,
        "■ 北海道ダム 地質複層推定可能性分析（所在地=北海道の全ダム）")

    hok = [d for d in dams if d["pref"].startswith("北海道")]
    row = _write_coverage_sections(ws, 2, hok)
    _write_dam_detail(ws, row, hok)


# ─── A2: 全国地質推定分析（北海道除く） ─────────────────────
def write_a2(wb, dams, glossary):
    ws = wb.create_sheet("A2_全国地質推定分析")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 1, 13,
        "■ 全国ダム（北海道除く）地質複層推定可能性分析")

    non_hok = [d for d in dams if not d["pref"].startswith("北海道")]
    row = _write_coverage_sections(ws, 2, non_hok)

    # ── セクション 4: 都道府県別カバレッジ（A2のみ） ──
    sec_title(ws, row, 1, 7, "▼ 都道府県別カバレッジ（カバー率昇順）")
    row += 1
    for ci, (t, w) in enumerate(zip(
        ["都道府県", "総ダム数", "データあり", "カバー率(%)", "平均層数", "推定良好(≥3層)", "推定不可(0層)"],
        [12, 8, 10, 12, 10, 12, 12]
    ), 1):
        hdr(ws, row, ci, t, w)
    row += 1

    pref_dams = defaultdict(list)
    for d in non_hok:
        pref_dams[d["pref"]].append(d)

    for pref, pdams in sorted(
        pref_dams.items(),
        key=lambda x: pct(sum(1 for d in x[1] if d["layer_count"] > 0), len(x[1]))
    ):
        total_p = len(pdams)
        with_data_p = sum(1 for d in pdams if d["layer_count"] > 0)
        good_p   = sum(1 for d in pdams if d["layer_count"] >= 3)
        no_data_p = sum(1 for d in pdams if d["layer_count"] == 0)
        avg_lc = avg([d["layer_count"] for d in pdams])
        cr = pct(with_data_p, total_p)
        fill = GRN_FILL if cr >= 80 else YEL_FILL if cr >= 40 else RED_FILL
        for ci, v in enumerate(
            [pref, total_p, with_data_p, cr, avg_lc, good_p, no_data_p], 1
        ):
            body(ws, row, ci, v, align=RIGHT if ci > 1 else LEFT, fill=fill)
        row += 1
    row += 1

    _write_dam_detail(ws, row, non_hok,
        section_label=f"▼ ダム別詳細一覧（層数降順・堤高降順、全{len(non_hok)}件）")


# ─── A3: 情報充実推奨100ダム ─────────────────────────────────
def _enrich_reason(d):
    lc = d["layer_count"]
    if lc == 0:
        return "全5層データ未取得 — 最優先充実対象"
    filled_str = "・".join(LAYER_SHORT[n] for n in sorted(d["filled_layers"]))
    missing = 5 - lc
    return f"取得済み: {filled_str}（{lc}/5層）— 残り{missing}層を充実"


def write_a3(wb, dams, glossary):
    ws = wb.create_sheet("A3_情報充実推奨100ダム")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 1, 13,
        "■ 情報充実推奨100ダム（北海道除く）— symbol空白・不足ダムの優先リスト")
    row = 2

    # ── 選定方針 ──
    sec_title(ws, row, 1, 13, "▼ 選定方針")
    row += 1
    for line in [
        "対象: 北海道以外のダム（北海道は別途地質情報拡充予定のため除外）",
        "優先スコア = 不足層数×10 ＋ 堤高ボーナス（height÷5、最大10pt）＋ 管理者ボーナス（国交省/北海道開発局=3pt、その他=1pt）",
        "  例: 0層・堤高100m・国交省 → 5×10 + 20（cap10） + 3 = 63pt",
        "地域分散制約: 同一都道府県内から最大4件（100件未満の場合は制限解除して補充）",
        "目的: symbol空白・少数ダムにフィールド調査・API再実行の優先度を与え、全国地質DBの品質を向上させる",
    ]:
        cell = ws.cell(row, 1, line)
        cell.font = BODY_FONT; cell.alignment = LEFT
        row += 1
    row += 1

    # スコアリング（4層以下のダムを対象）
    candidates_raw = [d for d in dams if not d["pref"].startswith("北海道") and d["layer_count"] < 5]
    for d in candidates_raw:
        missing = 5 - d["layer_count"]
        h_bonus = min((d["height"] or 0) / 5, 10)
        mgr_bonus = 3 if d["mgr_code"] in (1, 10) else 1
        d["_enrich_score"] = missing * 10 + h_bonus + mgr_bonus

    candidates = sorted(candidates_raw, key=lambda d: -d["_enrich_score"])

    selected = []
    pref_cnt = Counter()
    for d in candidates:
        if len(selected) >= 100: break
        if pref_cnt[d["pref"]] < 4:
            selected.append(d)
            pref_cnt[d["pref"]] += 1

    if len(selected) < 100:
        added = {id(d) for d in selected}
        for d in candidates:
            if len(selected) >= 100: break
            if id(d) not in added:
                selected.append(d)
                added.add(id(d))

    # ── 選定サマリー ──
    sec_title(ws, row, 1, 3, "▼ 選定サマリー")
    row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    for label, val in [
        ("選定ダム数",              len(selected)),
        ("0層（全データ未取得）",   sum(1 for d in selected if d["layer_count"] == 0)),
        ("1層",                     sum(1 for d in selected if d["layer_count"] == 1)),
        ("2層",                     sum(1 for d in selected if d["layer_count"] == 2)),
        ("3層",                     sum(1 for d in selected if d["layer_count"] == 3)),
        ("4層",                     sum(1 for d in selected if d["layer_count"] == 4)),
        ("都道府県数",              len(pref_cnt)),
        ("平均充填済み層数",        avg([d["layer_count"] for d in selected])),
        ("平均充実スコア",          avg([d["_enrich_score"] for d in selected])),
    ]:
        body(ws, row, 1, label, bold=True)
        body(ws, row, 2, val, align=RIGHT)
        row += 1
    row += 1

    # ── 推奨リスト ──
    sec_title(ws, row, 1, 13, f"▼ 推奨100ダム一覧（充実スコア降順）")
    row += 1
    hdrs   = ["順位", "ダム名", "都道府県", "堤高(m)", "管理者", "現在層数",
              "L1-symbol", "L2-symbol", "L3-symbol", "L4-symbol", "L5-symbol",
              "充実スコア", "推奨理由"]
    widths = [6, 16, 12, 8, 8, 8, 16, 16, 16, 16, 16, 10, 50]
    hdr_row = row
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        hdr(ws, row, ci, h, w)
    row += 1

    for rank, d in enumerate(selected, 1):
        fill = RED_FILL if d["layer_count"] == 0 else \
               YEL_FILL if d["layer_count"] <= 2 else \
               ALT_FILL if row % 2 == 0 else None
        layer_syms = [
            d["layers"][n]["symbol"] if d["layers"][n] and d["layers"][n].get("symbol") else ""
            for n in range(1, 6)
        ]
        vals = [rank, d["name"], d["pref"], d["height"], d["mgr_code"],
                d["layer_count"]] + layer_syms + [
                round(d["_enrich_score"], 1), _enrich_reason(d)]
        for ci, v in enumerate(vals, 1):
            body(ws, row, ci, v, align=RIGHT if ci in (1, 4, 5, 6, 12) else LEFT, fill=fill)
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A{hdr_row}:{scol(len(hdrs))}{hdr_row}"


# ─── A4: 管理者別カバレッジ ──────────────────────────────────
def write_a4(wb, dams, glossary):
    ws = wb.create_sheet("A4_管理者別カバレッジ")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 1, 10,
        "■ 管理者別 地質推定可能性（北海道除く）— 充実優先機関の特定")
    row = 2

    non_hok = [d for d in dams if not d["pref"].startswith("北海道")]
    mgr_dams = defaultdict(list)
    for d in non_hok:
        mgr_dams[d["mgr_code"]].append(d)

    # ── サマリー表（カバー率昇順） ──
    sec_title(ws, row, 1, 10, "▼ 管理者コード別 推定可能性サマリー（カバー率昇順）")
    row += 1
    hdrs = ["管理者コード", "総ダム数", "データあり", "カバー率(%)",
            "推定良好(≥3)", "推定限定的(1-2)", "推定不可(0)",
            "平均層数", "代表都道府県（上位3）", "代表symbol（上位3）"]
    widths = [14, 8, 10, 12, 12, 14, 12, 10, 30, 40]
    hdr_row = row
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        hdr(ws, row, ci, h, w)
    row += 1

    mgr_rows = []
    for code, mdams in mgr_dams.items():
        total_m  = len(mdams)
        with_data = sum(1 for d in mdams if d["layer_count"] > 0)
        cr       = pct(with_data, total_m)
        good     = sum(1 for d in mdams if d["layer_count"] >= 3)
        partial  = sum(1 for d in mdams if 1 <= d["layer_count"] <= 2)
        no_data  = sum(1 for d in mdams if d["layer_count"] == 0)
        avg_lc   = avg([d["layer_count"] for d in mdams])
        pref_top3 = "、".join(p for p, _ in Counter(d["pref"] for d in mdams).most_common(3))
        sym_top3  = "、".join(s for s, _ in Counter(
            r["symbol"] for d in mdams for r in d["recs"]
        ).most_common(3))
        mgr_rows.append((cr, code, total_m, with_data, good, partial, no_data, avg_lc, pref_top3, sym_top3))

    for cr, code, total_m, with_data, good, partial, no_data, avg_lc, pref_top3, sym_top3 in sorted(mgr_rows):
        fill = GRN_FILL if cr >= 80 else YEL_FILL if cr >= 40 else RED_FILL
        for ci, v in enumerate(
            [code, total_m, with_data, cr, good, partial, no_data, avg_lc, pref_top3, sym_top3], 1
        ):
            body(ws, row, ci, v, align=RIGHT if ci in (2, 3, 4, 5, 6, 7, 8) else LEFT, fill=fill)
        row += 1
    row += 1

    # ── カバー率最低の上位2管理者: 推定不可ダム一覧 ──
    worst2 = sorted(
        [(code, mdams) for code, mdams in mgr_dams.items()
         if sum(1 for d in mdams if d["layer_count"] == 0) >= 3],
        key=lambda x: pct(sum(1 for d in x[1] if d["layer_count"] > 0), len(x[1]))
    )[:2]

    for code, mdams in worst2:
        gap_dams = sorted(
            [d for d in mdams if d["layer_count"] == 0],
            key=lambda d: -(d["height"] or 0)
        )[:10]
        cr_m = pct(sum(1 for d in mdams if d["layer_count"] > 0), len(mdams))
        n_gap = sum(1 for d in mdams if d["layer_count"] == 0)
        sec_title(ws, row, 1, 7,
            f"▼ 管理者コード {code}（カバー率 {cr_m}%、推定不可 {n_gap}件）— 堤高上位10件")
        row += 1
        for ci, (t, w) in enumerate(zip(
            ["ダム名", "都道府県", "堤高(m)", "管理者", "層数", "推定可能性", "所在地"],
            [16, 12, 8, 8, 6, 14, 24]
        ), 1):
            sub(ws, row, ci, t, w)
        row += 1
        for d in gap_dams:
            for ci, v in enumerate(
                [d["name"], d["pref"], d["height"], d["mgr_code"],
                 d["layer_count"], estimability_label(d["layer_count"]),
                 d["loc"][:22] if d["loc"] else ""], 1
            ):
                body(ws, row, ci, v, align=RIGHT if ci in (3, 4, 5) else LEFT, fill=RED_FILL)
            row += 1
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A{hdr_row}:{scol(len(hdrs))}{hdr_row}"


# ─── A5: 都道府県×地質層 マトリクス ──────────────────────────
def write_a5(wb, dams, glossary):
    ws = wb.create_sheet("A5_都道府県×地質層マトリクス")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 1, 10,
        "■ 都道府県×地質層 カバレッジマトリクス（北海道除く）")
    row = 2

    non_hok = [d for d in dams if not d["pref"].startswith("北海道")]
    pref_dams = defaultdict(list)
    for d in non_hok:
        pref_dams[d["pref"]].append(d)

    sec_title(ws, row, 1, 10,
        "▼ 各セル = Layer充填ダム数（色: 赤=10%未満, 黄=10〜39%, 緑=40%以上 / 行全体色=データあり率）")
    row += 1

    hdrs = ["都道府県", "総ダム数",
            "L1: Pre-N", "L2: N-古", "L3: N-新", "L4: Q-old", "L5: Q-H",
            "平均層数", "データあり率(%)", "推定良好率(%)"]
    widths = [12, 8, 10, 10, 10, 10, 10, 10, 14, 14]
    hdr_row = row
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        hdr(ws, row, ci, h, w)
    row += 1

    grand_layer_counts = [0] * 5
    for pref, pdams in sorted(pref_dams.items(), key=lambda x: -len(x[1])):
        total_p   = len(pdams)
        lc_list   = [sum(1 for d in pdams if d["layers"][lnum] is not None) for lnum in range(1, 6)]
        avg_lc    = avg([d["layer_count"] for d in pdams])
        cr        = pct(sum(1 for d in pdams if d["layer_count"] > 0), total_p)
        good_pct  = pct(sum(1 for d in pdams if d["layer_count"] >= 3), total_p)
        row_fill  = GRN_FILL if cr >= 80 else YEL_FILL if cr >= 40 else RED_FILL

        body(ws, row, 1, pref,    align=LEFT,  fill=row_fill)
        body(ws, row, 2, total_p, align=RIGHT, fill=row_fill)
        for ci_off, lcount in enumerate(lc_list, 3):
            rate = pct(lcount, total_p)
            cell_fill = GRN_FILL if rate >= 40 else YEL_FILL if rate >= 10 else RED_FILL
            body(ws, row, ci_off, lcount, align=RIGHT, fill=cell_fill)
        body(ws, row,  8, avg_lc,   align=RIGHT, fill=row_fill)
        body(ws, row,  9, cr,       align=RIGHT, fill=row_fill)
        body(ws, row, 10, good_pct, align=RIGHT, fill=row_fill)

        for i, lc in enumerate(lc_list):
            grand_layer_counts[i] += lc
        row += 1

    # 合計行
    row += 1
    grand_total = len(non_hok)
    grand_cr    = pct(sum(1 for d in non_hok if d["layer_count"] > 0), grand_total)
    grand_good  = pct(sum(1 for d in non_hok if d["layer_count"] >= 3), grand_total)
    grand_avg   = avg([d["layer_count"] for d in non_hok])
    sub(ws, row, 1, "合計 / 平均", 12)
    sub(ws, row, 2, grand_total, 8)
    for ci_off, lc in enumerate(grand_layer_counts, 3):
        sub(ws, row, ci_off, lc, 10)
    sub(ws, row,  8, grand_avg,  10)
    sub(ws, row,  9, grand_cr,   14)
    sub(ws, row, 10, grand_good, 14)

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A{hdr_row}:{scol(len(hdrs))}{hdr_row}"


# ─── A6: 地質層別 空白補完分析 ────────────────────────────────
def write_a6(wb, dams, glossary):
    ws = wb.create_sheet("A6_層別空白補完分析")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 1, 9,
        "■ 地質層別 空白補完優先分析（北海道除く）— 各層の充実優先ダム TOP20")
    row = 2

    non_hok = [d for d in dams if not d["pref"].startswith("北海道")]
    total_non_hok = len(non_hok)

    # ── 層別 空白サマリー ──
    sec_title(ws, row, 1, 7, "▼ 層別 空白ダム数サマリー")
    row += 1
    for ci, (t, w) in enumerate(zip(
        ["層番号", "層名", "空白ダム数", "空白率(%)",
         "うち他層あり（補完可能性あり）", "全層空白（最低優先）", "充填済み代表symbol（上位3）"],
        [8, 30, 10, 10, 22, 18, 40]
    ), 1):
        hdr(ws, row, ci, t, w)
    row += 1

    for lnum in range(1, 6):
        missing   = [d for d in non_hok if d["layers"][lnum] is None]
        has_other = [d for d in missing  if d["layer_count"] > 0]
        zero_all  = [d for d in missing  if d["layer_count"] == 0]
        has_layer = [d for d in non_hok  if d["layers"][lnum] is not None]
        top_syms  = "、".join(s for s, _ in Counter(
            d["layers"][lnum]["symbol"]
            for d in has_layer
            if d["layers"][lnum] and d["layers"][lnum].get("symbol")
        ).most_common(3))
        miss_rate = pct(len(missing), total_non_hok)
        fill = RED_FILL if miss_rate >= 70 else YEL_FILL if miss_rate >= 40 else GRN_FILL
        for ci, v in enumerate(
            [lnum, LAYER_NAMES[lnum], len(missing), miss_rate,
             len(has_other), len(zero_all), top_syms], 1
        ):
            body(ws, row, ci, v, align=RIGHT if ci in (1, 3, 4, 5, 6) else LEFT, fill=fill)
        row += 1
    row += 1

    # ── 各層: 補完優先ダム TOP20 ──
    # 対象: 当該層が空白 かつ 他に1層以上あるダム（既存データが文脈になる）
    # スコア: 既存層数×5 + 堤高ボーナス(max10) + 管理者ボーナス
    for lnum in range(1, 6):
        candidates = [d for d in non_hok if d["layers"][lnum] is None and d["layer_count"] > 0]
        top20 = sorted(
            candidates,
            key=lambda d: (d["layer_count"] * 5
                           + min((d["height"] or 0) / 10, 10)
                           + (3 if d["mgr_code"] in (1, 10) else 1)),
            reverse=True
        )[:20]

        n_cand = len(candidates)
        sec_title(ws, row, 1, 9,
            f"▼ Layer{lnum}: {LAYER_NAMES[lnum]}"
            f" — 補完優先ダム TOP{min(20, n_cand)}件（対象 {n_cand}件中）")
        row += 1

        if not top20:
            cell = ws.cell(row, 1, "（対象ダムなし）")
            cell.font = BODY_FONT; cell.alignment = LEFT
            row += 2
            continue

        hdrs_l  = ["優先順", "ダム名", "都道府県", "堤高(m)", "管理者",
                   "現在層数", "取得済みLayer（既存symbols）",
                   "平均強度スコア", "平均透水性スコア"]
        widths_l = [8, 16, 12, 8, 8, 8, 52, 14, 14]
        for ci, (h, w) in enumerate(zip(hdrs_l, widths_l), 1):
            sub(ws, row, ci, h, w)
        row += 1

        for pri, d in enumerate(top20, 1):
            fill = ALT_FILL if row % 2 == 0 else None
            filled_str = "  ".join(
                f"L{n}:{d['layers'][n]['symbol']}"
                for n in sorted(d["filled_layers"])
                if d["layers"][n] and d["layers"][n].get("symbol")
            )
            b_avg = avg([b_score(r["bearing_cap"])   for r in d["recs"] if b_score(r["bearing_cap"])])
            p_avg = avg([p_score(r["permeability"]) for r in d["recs"] if p_score(r["permeability"])])
            vals = [pri, d["name"], d["pref"], d["height"], d["mgr_code"],
                    d["layer_count"], filled_str, b_avg, p_avg]
            for ci, v in enumerate(vals, 1):
                body(ws, row, ci, v, align=RIGHT if ci in (1, 4, 5, 6, 8, 9) else LEFT, fill=fill)
            row += 1
        row += 1

    ws.freeze_panes = "A3"


# ─── 入力ファイル自動検索 ─────────────────────────────────────
def find_input_file(data_dir="data"):
    data_path = Path(data_dir)
    if not data_path.exists():
        raise FileNotFoundError(f"dataフォルダが見つかりません: {data_path.resolve()}")
    candidates = sorted(data_path.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError("dataフォルダ内にxlsxファイルが見つかりません")
    for path in candidates:
        try:
            wb = load_workbook(path, read_only=True)
            sheets = wb.sheetnames; wb.close()
            if "全国ダム地質DB" in sheets and "Glossary" in sheets:
                return path
        except Exception:
            continue
    raise FileNotFoundError("dataフォルダ内に「全国ダム地質DB」と「Glossary」シートを持つxlsxが見つかりません")


def parse_args():
    p = argparse.ArgumentParser(
        description="全国ダム地質DB 地質推定可能性分析 v1",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "例:\n"
            "  python3 分析3.py                              # data/ を自動検索\n"
            "  python3 分析3.py --input data/myfile.xlsx     # ファイルを直接指定\n"
            "  python3 分析3.py --output results/分析3.xlsx  # 出力先を指定\n"
        )
    )
    p.add_argument("--input",  default=None, help="入力xlsx（省略時はdata/フォルダを自動検索）")
    p.add_argument("--output", default=None, help="出力xlsx（省略時はdata/分析3結果_YYYYMMDD_HHMM.xlsx）")
    p.add_argument("--data",   default="data", help="自動検索フォルダ（デフォルト: data）")
    return p.parse_args()


def main():
    args = parse_args()

    if args.input:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"エラー: 入力ファイルが見つかりません: {input_path}")
            sys.exit(1)
    else:
        print(f"{args.data}/ フォルダから入力ファイルを自動検索...")
        input_path = find_input_file(args.data)
        mtime = datetime.datetime.fromtimestamp(input_path.stat().st_mtime)
        print(f"  → 使用ファイル: {input_path.name}  (更新: {mtime.strftime('%Y-%m-%d %H:%M')})")

    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path(args.data) / f"分析3結果_{ts}.xlsx"

    print(f"読み込み: {input_path}")
    print(f"出力先:   {output_path}")

    wb_in = load_workbook(input_path)
    dams, glossary = load_data(wb_in)
    hok_count = sum(1 for d in dams if d["pref"].startswith("北海道"))
    has = sum(1 for d in dams if d["layer_count"] > 0)
    print(f"総ダム数: {len(dams)} / データあり: {has} / 北海道: {hok_count} / 北海道以外: {len(dams)-hok_count}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    print("A1 北海道地質推定分析...")
    write_a1(wb_out, dams, glossary)
    print("A2 全国地質推定分析...")
    write_a2(wb_out, dams, glossary)
    print("A3 情報充実推奨100ダム...")
    write_a3(wb_out, dams, glossary)
    print("A4 管理者別カバレッジ...")
    write_a4(wb_out, dams, glossary)
    print("A5 都道府県×地質層マトリクス...")
    write_a5(wb_out, dams, glossary)
    print("A6 層別空白補完分析...")
    write_a6(wb_out, dams, glossary)

    wb_out.save(output_path)
    print(f"\n保存完了: {output_path}")


if __name__ == "__main__":
    main()
