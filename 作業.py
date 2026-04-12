"""
全国ダム地質DB - GeoNAVI API 地質情報取得・書き込みスクリプト
=============================================================
依存ライブラリ: openpyxl のみ（urllib は Python 標準ライブラリ）

インストール:
    pip3 install openpyxl
    または: python3 -m pip install openpyxl

使い方:
    python3 作業.py --input 全国ダム地質DB.xlsx --output 出力.xlsx
    python3 作業.py --input 全国ダム地質DB.xlsx --output 出力.xlsx --rows 3-10 --dry-run
    python3 作業.py --input 全国ダム地質DB.xlsx --output 出力.xlsx --overwrite

オプション:
    --rows 3-100   処理行範囲（省略時=全行）
    --overwrite    既存の凡例IDを上書き（省略時はスキップ）
    --dry-run      書き込みなし・確認のみ
    --log FILE     ログCSVファイルパス（デフォルト: 作業ログ.csv）

【層の定義】
    cont-1  W列  (凡例ID-1) : Pre-N  新第三紀以前
    cont-2  AI列 (凡例ID-2) : N      新第三紀（古い方）
    cont-3  AU列 (凡例ID-3) : N      新第三紀（新しい方）
    cont-4  BG列 (凡例ID-4) : Q-old  中期更新世
    cont-5  BS列 (凡例ID-5) : Q-H    後期更新世〜完新世

【後期更新世ヒット時の周辺探索】
    ダム地点が Q-H のみの場合、8方位にオフセット点を配置して探索。
    初期半径 500m → 見つからなければ 2 倍ずつ拡大（最大 8000m）。
    Q-H 以外の地質が最多出現した凡例IDを採用。
"""

import argparse
import csv
import json
import math
import ssl
import sys
import time
import urllib.request
import urllib.parse
import urllib.error
from pathlib import Path

# Mac の Python は OS 証明書バンドルを自動参照しないため
# SSL 検証を無効化したコンテキストを用意する（公開APIへのアクセスのため問題なし）
_SSL_CTX = ssl.create_default_context()
_SSL_CTX.check_hostname = False
_SSL_CTX.verify_mode = ssl.CERT_NONE

from openpyxl import load_workbook

# ─────────────────────────────────────────────
# 定数
# ─────────────────────────────────────────────

GEONAVI_URL  = "https://gbank.gsj.jp/seamless/v2/api/1.2/legend.json"
API_INTERVAL = 0.5      # APIコール間隔(秒)
REQUEST_TIMEOUT = 15    # タイムアウト(秒)
MAX_RETRY    = 3        # リトライ回数
SEARCH_RADII = [500, 1000, 2000, 4000, 8000]   # 周辺探索半径(m)

# 8方位探索方向 (dx係数, dy係数)
SEARCH_DIRECTIONS = [
    ( 1,  0), (-1,  0), ( 0,  1), ( 0, -1),
    ( 1,  1), ( 1, -1), (-1,  1), (-1, -1),
]

# geo_era → 使用する cont 層番号リスト（N は最大2枠）
ERA_TO_LAYER = {
    "Pre-N": [1],
    "N":     [2, 3],
    "Q-old": [4],
    "Q-H":   [5],
}

# cont 層番号 → 凡例ID列番号（1-indexed）
LAYER_ID_COL = {
    1: 23,   # W列   凡例ID-1
    2: 35,   # AI列  凡例ID-2
    3: 47,   # AU列  凡例ID-3
    4: 59,   # BG列  凡例ID-4
    5: 71,   # BS列  凡例ID-5
}

# formationAge_ja の地質年代順ソートキー（古い順 = 小さい値）
FORMATION_AGE_ORDER = {
    "先カンブリア時代": 10,
    "古生代 カンブリア紀": 21, "古生代 オルドビス紀": 22, "古生代 シルル紀": 23,
    "古生代 デボン紀": 24,     "古生代 石炭紀": 25,       "古生代 ペルム紀": 26,
    "古生代": 20,
    "中生代 三畳紀": 31,       "中生代 ジュラ紀": 32,     "中生代 白亜紀": 33,
    "中生代": 30,
    "古第三紀 暁新世": 41,     "古第三紀 始新世": 42,     "古第三紀 漸新世": 43,
    "古第三紀": 40,
    "新第三紀 中新世": 51,     "新第三紀 鮮新世": 52,
    "新第三紀": 50,
    "第四紀 前期更新世": 62,   "第四紀 中期更新世": 63,   "第四紀 後期更新世": 64,
    "第四紀 完新世": 65,       "第四紀 更新世": 61,       "第四紀": 60,
}


def age_sort_key(age_str):
    """formationAge_ja から年代順ソートキーを返す（前方一致・長いキー優先）"""
    if not age_str:
        return 9999
    best = 9999
    for k, v in FORMATION_AGE_ORDER.items():
        if age_str.startswith(k) and v < best:
            best = v
    return best


# ─────────────────────────────────────────────
# Glossary 読み込み
# ─────────────────────────────────────────────

def load_glossary(wb):
    """
    Glossary シートを読み込み、以下2つの辞書を返す。
      by_id     : {凡例ID(int): {field: value}}
      by_symbol : {symbol(str): {field: value}}
    列順: id / symbol / geo_surface / geo_era / geo_rock /
          formationAge_ja / group_ja / lithology_ja / geo_rock_label /
          bearing_cap / permeability / main_risk
    """
    ws = wb["Glossary"]
    fields = [
        "id", "symbol", "geo_surface", "geo_era", "geo_rock",
        "formationAge_ja", "group_ja", "lithology_ja", "geo_rock_label",
        "bearing_cap", "permeability", "main_risk",
    ]
    by_id     = {}
    by_symbol = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        try:
            rec = {f: row[i] for i, f in enumerate(fields)}
            gid = int(rec["id"])
            by_id[gid] = rec
            if rec["symbol"]:
                by_symbol[str(rec["symbol"])] = rec
        except (TypeError, ValueError):
            continue
    return by_id, by_symbol


# ─────────────────────────────────────────────
# GeoNAVI API 呼び出し（urllib 標準ライブラリ使用）
# ─────────────────────────────────────────────

def call_geonavi(lat, lng):
    """
    GeoNAVI API を呼び出し、(ids, symbols) を返す。
    - point= 指定時: dict 1件 {"symbol": "...", ...}
    - box=  指定時: list [{...}, ...]
    失敗時は ([], [])。
    """
    url = f"{GEONAVI_URL}?point={lat:.8f},{lng:.8f}"
    for attempt in range(1, MAX_RETRY + 1):
        try:
            req = urllib.request.Request(
                url,
                headers={"User-Agent": "propylite-dam-geology/1.0"}
            )
            with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT, context=_SSL_CTX) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            # point= のレスポンスは dict 1件
            if isinstance(data, dict):
                items = [data]
            elif isinstance(data, list):
                items = data
            else:
                return [], [], False, False
            ids     = []
            symbols = []
            all_null = True   # 全アイテムが null symbol かどうか
            for item in items:
                if not isinstance(item, dict):
                    continue
                code = item.get("code") or item.get("legend_id") or item.get("id")
                if code is not None:
                    try:
                        ids.append(int(code))
                        all_null = False
                        continue
                    except (ValueError, TypeError):
                        pass
                sym = item.get("symbol")
                if sym:
                    symbols.append(str(sym))
                    all_null = False
            # all_null=True は symbol:null（地質データなし）
            return ids, symbols, all_null
        except urllib.error.HTTPError as e:
            print(f"\n    [WARN] HTTP {e.code} lat={lat} lng={lng}", file=sys.stderr)
            if attempt == MAX_RETRY:
                return [], [], False
            time.sleep(2 ** attempt)
        except Exception as e:
            if attempt == MAX_RETRY:
                print(f"\n    [WARN] API失敗 lat={lat} lng={lng}: {e}", file=sys.stderr)
                return [], [], False
            time.sleep(2 ** attempt)
    return [], [], False

def _resolve_ids(ids, symbols, by_symbol):
    """
    call_geonavi が返した ids (直接ID) と symbols (symbol文字列) を統合し、
    凡例IDのリストを返す。symbolはby_symbolで凡例IDに変換する。
    """
    result = list(ids)
    seen   = set(ids)
    for sym in symbols:
        rec = by_symbol.get(sym)
        if rec:
            gid = int(rec["id"])
            if gid not in seen:
                result.append(gid)
                seen.add(gid)
    return result


def offset_point(lat, lng, dx_m, dy_m):
    """中心座標から dx_m(東西), dy_m(南北) 移動した座標を返す"""
    dlat = dy_m / 111320.0
    dlng = dx_m / (111320.0 * math.cos(math.radians(lat)))
    return lat + dlat, lng + dlng


# ─────────────────────────────────────────────
# 後期更新世時の周辺探索
# ─────────────────────────────────────────────

def search_non_qh(lat, lng, glossary, by_symbol):
    """
    ダム地点が Q-H のみの場合、周辺8方位を探索して Q-H 以前の凡例IDを返す。

    探索手順:
      1. 初期半径 500m の8方位オフセット点で API を呼び出す
      2. Q-H 以外の凡例IDが得られれば採用候補とする
      3. 候補なし → 半径2倍で再探索（最大 8000m）
      4. 最多出現IDを優先リストの先頭に返す

    Returns:
      (found_ids: list, info: dict{"radius_m", "n_candidates"})
    """
    for radius in SEARCH_RADII:
        vote = {}   # {凡例ID: 出現数}
        for dx_f, dy_f in SEARCH_DIRECTIONS:
            p_lat, p_lng = offset_point(lat, lng, dx_f * radius, dy_f * radius)
            ids, syms, _ = call_geonavi(p_lat, p_lng)
            ids = _resolve_ids(ids, syms, by_symbol)
            time.sleep(API_INTERVAL)
            for gid in ids:
                rec = glossary.get(gid)
                if rec and rec.get("geo_era") != "Q-H":
                    vote[gid] = vote.get(gid, 0) + 1

        if vote:
            sorted_ids = [gid for gid, _ in sorted(vote.items(), key=lambda x: -x[1])]
            return sorted_ids, {"radius_m": radius, "n_candidates": len(vote)}

    return [], {"radius_m": SEARCH_RADII[-1], "n_candidates": 0}


# ─────────────────────────────────────────────
# 層割り当て
# ─────────────────────────────────────────────

def search_null_point(lat, lng, glossary, by_symbol):
    """
    symbol:null の地点で、周辺を探索して有効な地質を返す。
    探索方法:
      - 8方位 × 半径 500m から開始、見つからなければ 2倍ずつ拡大（最大 8000m）
      - null でない symbol が得られた地点の結果を採用
      - 複数方位でヒットした場合は出現数が最多の凡例IDを優先

    Returns:
      (found_ids: list, info: dict{"radius_m", "n_candidates"})
    """
    for radius in SEARCH_RADII:
        vote = {}   # {凡例ID: 出現数}
        for dx_f, dy_f in SEARCH_DIRECTIONS:
            p_lat, p_lng = offset_point(lat, lng, dx_f * radius, dy_f * radius)
            ids, syms, is_null = call_geonavi(p_lat, p_lng)
            time.sleep(API_INTERVAL)
            if is_null:
                continue
            ids = _resolve_ids(ids, syms, by_symbol)
            for gid in ids:
                if gid in glossary:
                    vote[gid] = vote.get(gid, 0) + 1

        if vote:
            sorted_ids = [gid for gid, _ in sorted(vote.items(), key=lambda x: -x[1])]
            return sorted_ids, {"radius_m": radius, "n_candidates": len(vote)}

    return [], {"radius_m": SEARCH_RADII[-1], "n_candidates": 0}


def assign_layers(valid_ids, glossary):
    """
    凡例IDリストを geo_era に基づき cont 層に割り当てる。
    N が複数ある場合は formationAge_ja の古い順に cont-2, cont-3 へ。

    Returns: {cont_number(int): 凡例ID(int)}
    """
    by_era = {"Pre-N": [], "N": [], "Q-old": [], "Q-H": []}
    seen = set()
    for gid in valid_ids:
        if gid in seen:
            continue
        seen.add(gid)
        rec = glossary.get(gid)
        if rec and rec.get("geo_era") in by_era:
            by_era[rec["geo_era"]].append(gid)

    result = {}

    # Pre-N → cont-1（複数あれば先頭のみ）
    if by_era["Pre-N"]:
        result[1] = by_era["Pre-N"][0]

    # N → cont-2, cont-3（古い順）
    if by_era["N"]:
        n_sorted = sorted(
            by_era["N"],
            key=lambda gid: age_sort_key(glossary[gid].get("formationAge_ja", ""))
        )
        for slot, gid in zip(ERA_TO_LAYER["N"], n_sorted):
            result[slot] = gid

    # Q-old → cont-4
    if by_era["Q-old"]:
        result[4] = by_era["Q-old"][0]

    # Q-H → cont-5
    if by_era["Q-H"]:
        result[5] = by_era["Q-H"][0]

    return result


# ─────────────────────────────────────────────
# Excel 書き込み
# ─────────────────────────────────────────────

def write_assignment(ws, row, assignment, glossary, overwrite, dry_run):
    """
    assignment {cont: 凡例ID} を Excel に書き込む。
    ・凡例ID列 (W/AI/AU/BG/BS) に ID を書き込む
    ・symbol〜main_risk 列に Glossary の値を直接転記する
    ・geo_comp (V列) を更新する
    既存の数式セルは保護しない（数式は全削除済みのため値を直接書く）。
    """
    # cont番号 → (凡例ID列, symbol列開始) のマッピング
    # 各contは: 凡例ID, symbol, geo_surface, geo_era, geo_rock,
    #           formationAge_ja, group_ja, lithology_ja, geo_rock_label,
    #           bearing_cap, permeability, main_risk の12列
    CONT_START = {
        1: 23,   # W
        2: 35,   # AI
        3: 47,   # AU
        4: 59,   # BG
        5: 71,   # BS
    }
    GLOSSARY_FIELDS = [
        "id", "symbol", "geo_surface", "geo_era", "geo_rock",
        "formationAge_ja", "group_ja", "lithology_ja", "geo_rock_label",
        "bearing_cap", "permeability", "main_risk",
    ]

    log = []
    for cont, gid in sorted(assignment.items()):
        start_col = CONT_START[cont]
        id_cell   = ws.cell(row=row, column=start_col)

        # 既存値チェック（上書きオプション）
        if id_cell.value is not None and not overwrite:
            log.append(f"cont-{cont}:skip(既存={id_cell.value})")
            continue

        rec = glossary.get(gid)
        if not rec:
            log.append(f"cont-{cont}:skip(Glossary未登録 id={gid})")
            continue

        if not dry_run:
            for i, field in enumerate(GLOSSARY_FIELDS):
                ws.cell(row=row, column=start_col + i).value = rec[field]

        log.append(f"cont-{cont}:id={gid}({rec.get('symbol','')})")

    # geo_comp (V列=22) を更新: geo_surface-1〜5 を \ 区切りで結合
    if not dry_run:
        surfaces = []
        for cont in [1, 2, 3, 4, 5]:
            gs_col = CONT_START[cont] + 2   # geo_surface は ID列+2
            v = ws.cell(row=row, column=gs_col).value
            if v:
                surfaces.append(str(v))
        ws.cell(row=row, column=22).value = "\\".join(surfaces) if surfaces else None

    return log


# ─────────────────────────────────────────────
# メイン
# ─────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="全国ダム地質DB GeoNAVI API 取得・書き込みスクリプト"
    )
    p.add_argument("--input",     required=True,          help="入力 xlsx")
    p.add_argument("--output",    required=True,          help="出力 xlsx")
    p.add_argument("--rows",      default=None,           help="処理行範囲 例: 3-100")
    p.add_argument("--overwrite", action="store_true",    help="既存凡例IDを上書き")
    p.add_argument("--dry-run",   action="store_true",    help="書き込みなし確認のみ")
    p.add_argument("--log",       default="作業ログ.csv", help="ログCSVパス")
    p.add_argument("--retry-log", default=None,           help="再処理するログCSV（api_errorの行のみ再実行）")
    return p.parse_args()


def parse_row_range(rows_str, max_row):
    if rows_str is None:
        return range(3, max_row + 1)
    parts = rows_str.split("-")
    start = int(parts[0])
    end   = int(parts[1]) if len(parts) == 2 else start
    return range(start, end + 1)


def load_retry_rows(log_path):
    """
    ログCSVから status が api_error の行番号セットを返す。
    """
    retry_rows = set()
    with open(log_path, newline="", encoding="utf-8-sig") as f:
        for rec in csv.DictReader(f):
            if rec.get("status") == "api_error":
                try:
                    retry_rows.add(int(rec["row"]))
                except (ValueError, KeyError):
                    pass
    return retry_rows


def main():
    args = parse_args()

    print("=== 全国ダム地質DB GeoNAVI 取得処理 ===")
    print(f"  入力     : {args.input}")
    print(f"  出力     : {args.output}")
    print(f"  上書き   : {args.overwrite}")
    print(f"  ドライラン: {args.dry_run}")

    wb       = load_workbook(args.input)
    ws       = wb["全国ダム地質DB"]
    glossary, by_symbol = load_glossary(wb)
    print(f"  Glossary : {len(glossary)} 件\n")

    # --retry-log 指定時はログのapi_error行のみ処理
    if args.retry_log:
        retry_rows = load_retry_rows(args.retry_log)
        row_list   = sorted(retry_rows)
        print(f"  再処理モード: {args.retry_log} の api_error {len(row_list)} 件\n")
    else:
        row_range = parse_row_range(args.rows, ws.max_row)
        row_list  = list(row_range)
        print(f"  処理行範囲: {row_range.start} 〜 {row_range.stop - 1}\n")

    log_records = []
    total = len(row_list)

    for i, row in enumerate(row_list, 1):
        dam_name = ws.cell(row=row, column=3).value
        lat      = ws.cell(row=row, column=20).value   # T列: 緯度
        lng      = ws.cell(row=row, column=21).value   # U列: 経度

        prefix = f"[{i:>4}/{total}] row{row} {dam_name or '(名称なし)'}"

        if not lat or not lng:
            print(f"{prefix}: 座標なし → スキップ")
            log_records.append({
                "row": row, "dam": dam_name, "lat": "", "lng": "",
                "status": "no_coords", "detail": "", "search_radius": "", "raw_ids": ""
            })
            continue

        print(f"{prefix} ({lat:.5f}, {lng:.5f}) ...", end=" ", flush=True)

        # ─── (1) ダム地点の地質取得 ───
        ids, syms, is_null = call_geonavi(lat, lng)
        ids = _resolve_ids(ids, syms, by_symbol)
        time.sleep(API_INTERVAL)

        if is_null:
            # symbol:null → 周辺を探索
            print("null→周辺探索...", end=" ", flush=True)
            ids, sinfo = search_null_point(lat, lng, glossary, by_symbol)
            search_radius = sinfo["radius_m"]
            if not ids:
                print(f"地質なし(r={search_radius}m)")
                log_records.append({
                    "row": row, "dam": dam_name, "lat": lat, "lng": lng,
                    "status": "no_data", "detail": f"周辺{search_radius}m圏内も地質なし",
                    "search_radius": search_radius, "raw_ids": ""
                })
                continue
            print(f"({len(ids)}件取得 r={search_radius}m)...", end=" ", flush=True)
            valid_ids = [gid for gid in ids if gid in glossary]
            assignment = assign_layers(valid_ids, glossary)
            written = write_assignment(ws, row, assignment, glossary, args.overwrite, args.dry_run)
            print(", ".join(written) if written else "変化なし")
            log_records.append({
                "row": row, "dam": dam_name, "lat": lat, "lng": lng,
                "status": "ok",
                "detail": "; ".join(written),
                "search_radius": search_radius,
                "raw_ids": str(ids),
            })
            continue

        if not ids:
            print("API失敗")
            log_records.append({
                "row": row, "dam": dam_name, "lat": lat, "lng": lng,
                "status": "api_error", "detail": "", "search_radius": "", "raw_ids": ""
            })
            continue

        # Glossary 照合
        valid_ids   = [gid for gid in ids if gid in glossary]
        unknown_ids = [gid for gid in ids if gid not in glossary]
        if unknown_ids:
            print(f"\n    [WARN] Glossary未登録: {unknown_ids}", file=sys.stderr)

        # ─── (2) Q-H のみ → 周辺探索 ───
        search_radius = ""
        eras = {glossary[gid]["geo_era"] for gid in valid_ids}

        if eras and eras <= {"Q-H"}:
            print("Q-Hのみ→周辺探索...", end=" ", flush=True)
            extra_ids, sinfo = search_non_qh(lat, lng, glossary, by_symbol)
            search_radius = sinfo["radius_m"]
            if extra_ids:
                seen = set(valid_ids)
                for gid in extra_ids:
                    if gid not in seen:
                        valid_ids.append(gid)
                        seen.add(gid)
                print(f"({len(extra_ids)}件追加 r={search_radius}m)...", end=" ", flush=True)
            else:
                print(f"(非Q-H地質なし r={search_radius}m)...", end=" ", flush=True)

        # ─── (3) 層割り当て ───
        assignment = assign_layers(valid_ids, glossary)

        # ─── (4) 書き込み ───
        written = write_assignment(ws, row, assignment, glossary, args.overwrite, args.dry_run)
        print(", ".join(written) if written else "変化なし")

        log_records.append({
            "row": row, "dam": dam_name, "lat": lat, "lng": lng,
            "status": "ok",
            "detail": "; ".join(written),
            "search_radius": search_radius,
            "raw_ids": str(ids),
        })

    # ─── 保存 ───
    if not args.dry_run:
        wb.save(args.output)
        print(f"\n保存完了: {args.output}")
    else:
        print(f"\n[ドライラン] 保存スキップ")

    # ─── ログ出力 ───
    fieldnames = ["row", "dam", "lat", "lng", "status", "detail", "search_radius", "raw_ids"]
    with open(args.log, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(log_records)
    print(f"ログ: {args.log}")

    ok      = sum(1 for r in log_records if r["status"] == "ok")
    no_data = sum(1 for r in log_records if r["status"] == "no_data")
    err     = sum(1 for r in log_records if r["status"] == "api_error")
    skip    = sum(1 for r in log_records if r["status"] == "no_coords")
    print(f"\n完了: 成功 {ok} 件 / データなし {no_data} 件 / 通信失敗 {err} 件 / 座標なし {skip} 件")


if __name__ == "__main__":
    main()
