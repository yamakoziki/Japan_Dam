"""
全国ダム地質DB 統計分析スクリプト v2
======================================
Symbol の1項目・2項目・3項目（地質時代_岩種_修飾子）の類似性に着目し、
強度（bearing_cap）と透水性（permeability）を中心に分析する。

北海道60ダム = 北海道×国土交通省管理（74件中、データあり上位60件）
全国100ダム  = 北海道を除く全国から選定（北海道ダムは含まない）

追加シート:
  S1_Symbol階層分析   : 1項目・2項目・3項目別の出現頻度と強度・透水性プロファイル
  S2_強度透水性マトリクス : bearing_cap × permeability の地質マトリクス
  S3_Symbol類似グループ  : 1項目共通Symbolの強度・透水性クラスター
  S4_2項目組合せ      : 2Symbol組合せ × 強度・透水性
  S5_北海道60ダム     : 北海道×国交省ダムのSymbol構成・強度透水性分析
  S6_全国100ダム選定  : 北海道除く全国から地質多様性・強度透水性分散を考慮した選定
  S7_カバレッジ比較   : 北海道60×全国100のSymbolカバレッジギャップ分析
"""

import argparse
import re
from collections import Counter, defaultdict
from itertools import combinations
from statistics import mean, stdev

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

# ─── スタイル ───────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUB_FILL  = PatternFill("solid", fgColor="2E75B6")
SUB_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SEC_FILL  = PatternFill("solid", fgColor="D6E4F0")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
ALT_FILL  = PatternFill("solid", fgColor="F2F7FC")
RED_FILL  = PatternFill("solid", fgColor="FFE0E0")
YEL_FILL  = PatternFill("solid", fgColor="FFF8DC")
GRN_FILL  = PatternFill("solid", fgColor="E8F5E9")
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT= Alignment(horizontal="right",  vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── 強度・透水性スコア ─────────────────────────────────────
BEARING_SCORE = {
    "低": 1, "低〜中": 2, "中": 3,
    "中〜高": 4, "中〜高（溶結度依存）": 4,
    "高（続成固結）": 5, "高": 5,
}
PERM_SCORE = {
    "低〜中": 1, "中（節理閉鎖）": 2, "中": 2, "中〜高": 3,
    "高": 4, "高（柱状節理）": 4, "高（柱状節理・カルデラ）": 4,
}
BEARING_ORDER = ["低","低〜中","中","中〜高","中〜高（溶結度依存）","高（続成固結）","高"]
PERM_ORDER    = ["低〜中","中（節理閉鎖）","中","中〜高","高","高（柱状節理）","高（柱状節理・カルデラ）"]
ERA_JA  = {"Pre-N":"新第三紀以前","N":"新第三紀","Q-old":"中期更新世","Q-H":"後期更新世〜完新世"}
ROCK_JA = {"UC":"未固結","SD":"堆積岩","PF":"火山岩","GR":"深成岩",
           "VL":"火砕岩","LS":"石灰岩","TC":"凝灰岩","MT":"変成岩","UM":"超苦鉄質岩"}

def b_score(v): return BEARING_SCORE.get(v, 0)
def p_score(v): return PERM_SCORE.get(v, 0)

# リスクランク: 支持力低×透水性高 = 最高リスク
def risk_rank(bearing, perm):
    b = b_score(bearing); p = p_score(perm)
    # 支持力低い（スコア小）かつ透水性高い（スコア大）ほど高リスク
    return (5 - b) + p  # 最大 = 4+4=8, 最小 = 0+1=1

def hdr(ws, r, c, v, w=None):
    cell = ws.cell(r, c, v)
    cell.font=HDR_FONT; cell.fill=HDR_FILL; cell.alignment=CTR; cell.border=BDR
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cell

def sub(ws, r, c, v, w=None):
    cell = ws.cell(r, c, v)
    cell.font=SUB_FONT; cell.fill=SUB_FILL; cell.alignment=CTR; cell.border=BDR
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cell

def body(ws, r, c, v, align=LEFT, bold=False, fill=None):
    cell = ws.cell(r, c, v)
    cell.font=BOLD_FONT if bold else BODY_FONT
    cell.alignment=align; cell.border=BDR
    if fill: cell.fill=fill
    return cell

def sec_title(ws, r, c1, c2, text):
    ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
    cell = ws.cell(r, c1, text)
    cell.font=BOLD_FONT; cell.fill=SEC_FILL; cell.alignment=LEFT
    return cell

def sheet_title(ws, r, c1, c2, text):
    ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
    cell = ws.cell(r, c1, text)
    cell.font=Font(name="Arial", bold=True, size=12, color="1F4E79")
    cell.alignment=LEFT
    return cell

def pct(n, total): return round(n/total*100,1) if total else 0
def avg(lst): return round(mean(lst),2) if lst else ""
def scol(c): return get_column_letter(c)

# ─── Symbol 分解 ────────────────────────────────────────────
def sym_parts(symbol):
    """
    例: J2-31_som_J2 → parts[0]="J2-31", parts[1]="som", parts[2]="J2"
    1項目=parts[0], 2項目=parts[0]+"_"+parts[1], 3項目=フル
    """
    parts = symbol.split("_")
    p1 = parts[0]
    p2 = "_".join(parts[:2]) if len(parts) >= 2 else symbol
    p3 = symbol
    return p1, p2, p3

# ─── データ読み込み ──────────────────────────────────────────
def load_data(wb):
    ws_db = wb["全国ダム地質DB"]
    ws_gl = wb["Glossary"]

    glossary = {}
    fields = ["id","symbol","geo_surface","geo_era","geo_rock",
              "formationAge_ja","group_ja","lithology_ja","geo_rock_label",
              "bearing_cap","permeability","main_risk"]
    for row in ws_gl.iter_rows(min_row=4, values_only=True):
        if row[0] is None: continue
        try:
            rec = {f: row[i] for i, f in enumerate(fields)}
            glossary[int(rec["id"])] = rec
        except: continue

    ID_COLS = {1:23, 2:35, 3:47, 4:59, 5:71}
    dams = []
    for row in range(3, ws_db.max_row+1):
        name = ws_db.cell(row, 3).value
        if not name: continue
        recs = []
        for cont in [1,2,3,4,5]:
            gid = ws_db.cell(row, ID_COLS[cont]).value
            if gid is not None and not isinstance(gid, str):
                rec = glossary.get(int(gid))
                if rec: recs.append(rec)
        loc  = ws_db.cell(row, 17).value or ""
        mgr  = ws_db.cell(row, 14).value
        height = ws_db.cell(row, 10).value
        dams.append({
            "row": row, "name": name, "loc": loc,
            "pref": loc[:3] if loc else "",
            "mgr_code": mgr,
            "height": height if isinstance(height, (int,float)) else None,
            "recs": recs,
        })
    return dams, glossary

# ─── S1: Symbol階層分析 ─────────────────────────────────────
def write_s1(wb, dams, glossary):
    ws = wb.create_sheet("S1_Symbol階層分析")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 1, 14, "■ Symbol 階層分析（1項目・2項目・3項目）— 強度・透水性プロファイル")

    for level, label, col_offset in [(1,"1項目（地質時代）",0),(2,"2項目（時代_岩種）",0),(3,"3項目（フルSymbol）",0)]:
        # 集計
        p_dams   = defaultdict(set)    # pN → ダムindexセット
        p_scores = defaultdict(list)   # pN → [(b_score, p_score), ...]
        p_risk   = defaultdict(list)
        p_recs   = {}                  # pN → 代表rec

        for di, d in enumerate(dams):
            for rec in d["recs"]:
                p1,p2,p3 = sym_parts(rec["symbol"])
                key = [p1,p2,p3][level-1]
                p_dams[key].add(di)
                bs = b_score(rec["bearing_cap"])
                ps = p_score(rec["permeability"])
                if bs and ps:
                    p_scores[key].append((bs, ps))
                    p_risk[key].append(risk_rank(rec["bearing_cap"], rec["permeability"]))
                if key not in p_recs:
                    p_recs[key] = rec

        # ヘッダー
        row = (level-1)*2 + 2 if level==1 else row+2
        # セクションタイトル
        sec_title(ws, row, 1, 14, f"▼ {label}（{len(p_dams)}種）")
        row += 1
        hdrs = [f"{label}コード","geo_era代表","geo_rock代表",
                "ダム数","延べ層数",
                "平均強度スコア(1-5)","平均透水性スコア(1-4)",
                "平均リスクスコア","強度低(%)","透水性高(%)","代表formationAge"]
        ws_widths = [20,10,8,8,8,16,16,14,10,10,28]
        for ci,(h,w) in enumerate(zip(hdrs,ws_widths),1):
            hdr(ws, row, ci, h, w)
        row += 1

        total_dams = len([d for d in dams if d["recs"]])
        for key, dam_set in sorted(p_dams.items(), key=lambda x: -len(x[1])):
            rec = p_recs.get(key, {})
            scores = p_scores[key]
            risks  = p_risk[key]
            b_avg  = avg([s[0] for s in scores])
            p_avg  = avg([s[1] for s in scores])
            r_avg  = avg(risks)
            b_low  = pct(sum(1 for s in scores if s[0]<=2), len(scores)) if scores else ""
            p_high = pct(sum(1 for s in scores if s[1]>=4), len(scores)) if scores else ""
            fill = ALT_FILL if row%2==0 else None
            # リスク高は赤、中は黄
            if isinstance(r_avg, float) and r_avg >= 6:
                fill = RED_FILL
            elif isinstance(r_avg, float) and r_avg >= 4:
                fill = YEL_FILL
            vals = [key, rec.get("geo_era",""), rec.get("geo_rock",""),
                    len(dam_set), sum(len(p_dams[key]) for _ in [key]),
                    b_avg, p_avg, r_avg, b_low, p_high,
                    rec.get("formationAge_ja","")]
            # 延べ層数は別途
            vals[4] = sum(1 for d in dams for r in d["recs"]
                         if sym_parts(r["symbol"])[level-1]==key)
            for ci,v in enumerate(vals,1):
                al = RIGHT if ci in (4,5,6,7,8,9,10) else LEFT
                body(ws, row, ci, v, align=al, fill=fill)
            row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N2"


# ─── S2: 強度×透水性マトリクス ──────────────────────────────
def write_s2(wb, dams, glossary):
    ws = wb.create_sheet("S2_強度透水性マトリクス")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, 1, 1, 12, "■ 強度（bearing_cap）× 透水性（permeability）地質マトリクス")

    # bearing × perm → {symbol: count, dams: set}
    matrix = defaultdict(lambda: {"count":0, "dams":set(), "syms":Counter(), "risks":[]})
    for di, d in enumerate(dams):
        for rec in d["recs"]:
            key = (rec["bearing_cap"], rec["permeability"])
            matrix[key]["count"] += 1
            matrix[key]["dams"].add(di)
            matrix[key]["syms"][rec["symbol"]] += 1
            matrix[key]["risks"].append(rec["main_risk"])

    # 実際に存在する値のみ
    bearings_exist = [b for b in BEARING_ORDER if any(k[0]==b for k in matrix)]
    perms_exist    = [p for p in PERM_ORDER    if any(k[1]==p for k in matrix)]

    row = 2
    # クロス集計（ダム数）
    sec_title(ws, row, 1, len(perms_exist)+3, "▼ ダム層数クロス集計（赤=高リスク 黄=中リスク）")
    row += 1
    ws.cell(row,1,"支持力↓ / 透水性→").fill=HDR_FILL
    ws.cell(row,1).font=HDR_FONT; ws.cell(row,1).alignment=CTR; ws.cell(row,1).border=BDR
    ws.column_dimensions["A"].width = 22
    for ci,p in enumerate(perms_exist,2):
        hdr(ws, row, ci, p, 16)
    hdr(ws, row, len(perms_exist)+2, "合計", 8)
    hdr(ws, row, len(perms_exist)+3, "リスクスコア", 12)
    row += 1

    for b in bearings_exist:
        bs = b_score(b)
        row_total = 0
        risk_scores = []
        body(ws,row,1,b,align=LEFT,bold=True,fill=PatternFill("solid",fgColor="D6E4F0"))
        for ci,p in enumerate(perms_exist,2):
            cnt  = matrix[(b,p)]["count"]
            rr   = risk_rank(b,p)
            fill_c = RED_FILL if rr>=6 else YEL_FILL if rr>=4 else GRN_FILL if cnt>0 else None
            body(ws,row,ci,cnt if cnt else "",align=RIGHT,fill=fill_c)
            row_total += cnt
            if cnt: risk_scores.append(rr)
        body(ws,row,len(perms_exist)+2,row_total,align=RIGHT,bold=True)
        avg_risk = avg(risk_scores)
        rfill = RED_FILL if isinstance(avg_risk,float) and avg_risk>=6 else \
                YEL_FILL if isinstance(avg_risk,float) and avg_risk>=4 else None
        body(ws,row,len(perms_exist)+3,avg_risk,align=RIGHT,fill=rfill)
        row += 1

    row += 2
    # 各セルの代表Symbol
    sec_title(ws, row, 1, 6, "▼ 強度×透水性 代表Symbol一覧（各セル上位3種）")
    row += 1
    for ci,t in enumerate(["支持力","透水性","リスクスコア","代表Symbol（上位3）","ダム数","ダム層数"],1):
        sub(ws,row,ci,t,22 if ci==4 else 14)
    row += 1

    for (b,p), data in sorted(matrix.items(),
                               key=lambda x: -risk_rank(x[0][0],x[0][1])):
        fill = RED_FILL if risk_rank(b,p)>=6 else YEL_FILL if risk_rank(b,p)>=4 else \
               ALT_FILL if row%2==0 else None
        top3 = "、".join(s for s,_ in data["syms"].most_common(3))
        vals = [b, p, risk_rank(b,p), top3, len(data["dams"]), data["count"]]
        for ci,v in enumerate(vals,1):
            al = RIGHT if ci in (3,5,6) else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row += 1

    ws.freeze_panes = "B4"


# ─── S3: Symbol類似グループ（1項目共通） ────────────────────
def write_s3(wb, dams, glossary):
    ws = wb.create_sheet("S3_Symbol類似グループ")
    ws.sheet_view.showGridLines = False
    sheet_title(ws,1,1,12,"■ Symbol類似グループ（1項目共通 = 地質時代コード一致）— 強度・透水性クラスター")

    # 1項目でグループ化
    groups = defaultdict(list)   # p1 → [rec, ...]
    group_dams = defaultdict(set)
    for di, d in enumerate(dams):
        for rec in d["recs"]:
            p1 = sym_parts(rec["symbol"])[0]
            groups[p1].append(rec)
            group_dams[p1].add(di)

    row = 2
    hdrs = ["1項目(地質時代)","含まれるSymbol数","ダム数","延べ層数",
            "平均強度スコア","強度最小","強度最大",
            "平均透水性スコア","透水性最小","透水性最大",
            "平均リスクスコア","主リスク(最多)","Symbol一覧(上位5)"]
    widths = [16,12,8,8,14,8,8,14,10,10,12,20,45]
    for ci,(h,w) in enumerate(zip(hdrs,widths),1):
        hdr(ws,row,ci,h,w)
    row += 1

    for p1, recs in sorted(groups.items(), key=lambda x: -len(group_dams[x[0]])):
        b_scores = [b_score(r["bearing_cap"]) for r in recs if b_score(r["bearing_cap"])]
        p_scores = [p_score(r["permeability"]) for r in recs if p_score(r["permeability"])]
        risks    = [risk_rank(r["bearing_cap"],r["permeability"]) for r in recs]
        risk_top = Counter(r["main_risk"] for r in recs).most_common(1)
        sym_top5 = "、".join(s for s,_ in Counter(r["symbol"] for r in recs).most_common(5))
        unique_syms = len(set(r["symbol"] for r in recs))

        r_avg = avg(risks)
        fill = RED_FILL if isinstance(r_avg,float) and r_avg>=6 else \
               YEL_FILL if isinstance(r_avg,float) and r_avg>=4 else \
               ALT_FILL if row%2==0 else None

        vals = [p1, unique_syms, len(group_dams[p1]), len(recs),
                avg(b_scores), min(b_scores) if b_scores else "",
                max(b_scores) if b_scores else "",
                avg(p_scores), min(p_scores) if p_scores else "",
                max(p_scores) if p_scores else "",
                r_avg, risk_top[0][0] if risk_top else "", sym_top5]
        for ci,v in enumerate(vals,1):
            al = RIGHT if ci in range(2,13) else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{scol(len(hdrs))}2"


# ─── S4: 2項目組合せ × 強度透水性 ─────────────────────────
def write_s4(wb, dams, glossary):
    ws = wb.create_sheet("S4_2項目組合せ")
    ws.sheet_view.showGridLines = False
    sheet_title(ws,1,1,14,"■ Symbol 2項目組合せ × 強度・透水性分析")

    combo_data = defaultdict(lambda:{
        "dams":set(),"b_scores":[],"p_scores":[],"risks":[],"era":[]})

    for di, d in enumerate(dams):
        syms = [(r["symbol"],r) for r in d["recs"]]
        seen = set()
        for (sa,ra),(sb,rb) in combinations(syms,2):
            key = tuple(sorted([sa,sb]))
            if key in seen: continue
            seen.add(key)
            combo_data[key]["dams"].add(di)
            for r in [ra,rb]:
                bs=b_score(r["bearing_cap"]); ps=p_score(r["permeability"])
                if bs: combo_data[key]["b_scores"].append(bs)
                if ps: combo_data[key]["p_scores"].append(ps)
                combo_data[key]["risks"].append(risk_rank(r["bearing_cap"],r["permeability"]))
                combo_data[key]["era"].append(r["geo_era"])

    row = 2
    hdrs = ["Symbol-A","1項目-A","Symbol-B","1項目-B",
            "地質時代構成","ダム数","平均強度","平均透水性","平均リスク",
            "強度最小","透水性最大","主リスク-A","主リスク-B"]
    widths = [20,12,20,12,30,8,10,10,10,10,10,20,20]
    for ci,(h,w) in enumerate(zip(hdrs,widths),1):
        hdr(ws,row,ci,h,w)
    row += 1

    for (sa,sb), data in sorted(combo_data.items(), key=lambda x:-len(x[1]["dams"]))[:200]:
        ra = next((r for r in glossary.values() if r["symbol"]==sa),{})
        rb = next((r for r in glossary.values() if r["symbol"]==sb),{})
        r_avg = avg(data["risks"])
        fill = RED_FILL if isinstance(r_avg,float) and r_avg>=6 else \
               YEL_FILL if isinstance(r_avg,float) and r_avg>=4 else \
               ALT_FILL if row%2==0 else None
        era_pair = f"{ERA_JA.get(ra.get('geo_era',''),'')} + {ERA_JA.get(rb.get('geo_era',''),'')}"
        vals=[sa, sym_parts(sa)[0], sb, sym_parts(sb)[0], era_pair,
              len(data["dams"]), avg(data["b_scores"]), avg(data["p_scores"]), r_avg,
              min(data["b_scores"]) if data["b_scores"] else "",
              max(data["p_scores"]) if data["p_scores"] else "",
              ra.get("main_risk",""), rb.get("main_risk","")]
        for ci,v in enumerate(vals,1):
            al=RIGHT if ci in (6,7,8,9,10,11) else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row += 1

    ws.freeze_panes="A3"
    ws.auto_filter.ref=f"A2:{scol(len(hdrs))}2"


# ─── S5: 北海道60ダム ───────────────────────────────────────
def write_s5(wb, dams, glossary):
    ws = wb.create_sheet("S5_北海道60ダム")
    ws.sheet_view.showGridLines = False
    sheet_title(ws,1,1,14,"■ 北海道60ダム（国土交通省管理）— Symbol構成・強度透水性分析")

    # 北海道×国交省(mgr_code=1)、データありを堤高降順で60件
    hok = [d for d in dams
           if d["pref"].startswith("北海道") and d["mgr_code"]==1 and d["recs"]]
    hok_sorted = sorted(hok, key=lambda d: -(d["height"] or 0))
    hok60 = hok_sorted[:60]

    row = 2
    # サマリー
    all_syms_hok = [r["symbol"] for d in hok60 for r in d["recs"]]
    p1_hok = [sym_parts(s)[0] for s in all_syms_hok]
    p2_hok = [sym_parts(s)[1] for s in all_syms_hok]
    b_scores_hok = [b_score(r["bearing_cap"]) for d in hok60
                    for r in d["recs"] if b_score(r["bearing_cap"])]
    p_scores_hok = [p_score(r["permeability"]) for d in hok60
                    for r in d["recs"] if p_score(r["permeability"])]

    summary = [
        ("対象ダム数（国交省×北海道）", len(hok)),
        ("選定ダム数（堤高上位60）", len(hok60)),
        ("ユニークSymbol数（3項目）", len(set(all_syms_hok))),
        ("ユニーク1項目数", len(set(p1_hok))),
        ("ユニーク2項目数", len(set(p2_hok))),
        ("平均強度スコア(1-5)", avg(b_scores_hok)),
        ("平均透水性スコア(1-4)", avg(p_scores_hok)),
        ("強度低（スコア≤2）ダム層数(%)", pct(sum(1 for s in b_scores_hok if s<=2),len(b_scores_hok))),
        ("透水性高（スコア≥4）ダム層数(%)", pct(sum(1 for s in p_scores_hok if s>=4),len(p_scores_hok))),
    ]
    ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=16
    for label,val in summary:
        ws.cell(row,1,label).font=BOLD_FONT; ws.cell(row,1).alignment=LEFT
        ws.cell(row,2,val).font=BODY_FONT;  ws.cell(row,2).alignment=RIGHT
        row += 1
    row += 1

    # ダム一覧
    sec_title(ws,row,1,14,"▼ 選定60ダム一覧（堤高降順）")
    row += 1
    hdrs=["ダム名","堤高(m)","所在地","Symbol-1","1項目-1","2項目-1",
          "Symbol-2","Symbol-3","Symbol-4","Symbol-5",
          "平均強度","平均透水性","平均リスク","主リスク"]
    widths=[16,8,20,18,12,16,16,16,16,16,10,10,10,22]
    for ci,(h,w) in enumerate(zip(hdrs,widths),1): hdr(ws,row,ci,h,w)
    row += 1

    for d in hok60:
        syms=[r["symbol"] for r in d["recs"]]
        bs=[b_score(r["bearing_cap"]) for r in d["recs"] if b_score(r["bearing_cap"])]
        ps=[p_score(r["permeability"]) for r in d["recs"] if p_score(r["permeability"])]
        risks=[risk_rank(r["bearing_cap"],r["permeability"]) for r in d["recs"]]
        risk_top=Counter(r["main_risk"] for r in d["recs"]).most_common(1)
        fill=ALT_FILL if row%2==0 else None
        s1=syms[0] if len(syms)>0 else ""
        vals=[d["name"],d["height"],d["loc"][:20] if d["loc"] else "",
              s1, sym_parts(s1)[0] if s1 else "", sym_parts(s1)[1] if s1 else "",
              syms[1] if len(syms)>1 else "",
              syms[2] if len(syms)>2 else "",
              syms[3] if len(syms)>3 else "",
              syms[4] if len(syms)>4 else "",
              avg(bs),avg(ps),avg(risks),
              risk_top[0][0] if risk_top else ""]
        for ci,v in enumerate(vals,1):
            al=RIGHT if ci in (2,11,12,13) else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row += 1

    row += 2
    # Symbol頻度
    sec_title(ws,row,1,8,"▼ 北海道60ダム Symbol頻度（1項目・2項目・3項目）")
    row += 1
    for ci,t in enumerate(["1項目","件数","2項目","件数","3項目フル","件数","強度avg","透水性avg"],1):
        sub(ws,row,ci,t,18 if ci in(1,3,5) else 8)
    row += 1
    p1c=Counter(p1_hok); p2c=Counter(p2_hok); p3c=Counter(all_syms_hok)
    p1_list=p1c.most_common(); p2_list=p2c.most_common(); p3_list=p3c.most_common()
    max_len=max(len(p1_list),len(p2_list),len(p3_list))
    for i in range(max_len):
        fill=ALT_FILL if row%2==0 else None
        v1=p1_list[i] if i<len(p1_list) else ("","")
        v2=p2_list[i] if i<len(p2_list) else ("","")
        v3=p3_list[i] if i<len(p3_list) else ("","")
        # 強度・透水性
        if v3[0]:
            rec=next((r for r in glossary.values() if r["symbol"]==v3[0]),{})
            b_a=b_score(rec.get("bearing_cap","")); p_a=p_score(rec.get("permeability",""))
        else: b_a=p_a=""
        for ci,v in enumerate([v1[0],v1[1],v2[0],v2[1],v3[0],v3[1],b_a,p_a],1):
            al=RIGHT if ci in(2,4,6,7,8) else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row += 1

    ws.freeze_panes="A3"


# ─── S6: 全国100ダム選定 ─────────────────────────────────────
def write_s6(wb, dams, glossary):
    ws = wb.create_sheet("S6_全国100ダム選定")
    ws.sheet_view.showGridLines = False
    sheet_title(ws,1,1,16,"■ 全国100ダム選定（北海道除く）— 地質多様性・強度透水性分散を考慮")

    # 北海道除外
    non_hok = [d for d in dams
               if not d["pref"].startswith("北海道") and d["recs"]]

    # 北海道60の既カバーSymbol（p1・p2・p3）
    hok = [d for d in dams
           if d["pref"].startswith("北海道") and d["mgr_code"]==1 and d["recs"]]
    hok60 = sorted(hok, key=lambda d:-(d["height"] or 0))[:60]
    hok_p1=set(sym_parts(r["symbol"])[0] for d in hok60 for r in d["recs"])
    hok_p2=set(sym_parts(r["symbol"])[1] for d in hok60 for r in d["recs"])
    hok_p3=set(r["symbol"]              for d in hok60 for r in d["recs"])

    # 全Symbolの全国頻度
    all_sym_cnt=Counter(r["symbol"] for d in dams for r in d["recs"])

    # スコアリング: 希少Symbol保有 + 北海道未カバーSymbol保有 + 堤高分散
    def score_dam(d):
        syms=[r["symbol"] for r in d["recs"]]
        p1s=[sym_parts(s)[0] for s in syms]
        p2s=[sym_parts(s)[1] for s in syms]
        # 希少性スコア（全国出現少ないほど高得点）
        rarity = sum(1.0/all_sym_cnt[s] for s in syms)
        # 北海道未カバーP1数（新しい地質時代）
        new_p1 = sum(1 for p in p1s if p not in hok_p1)
        new_p2 = sum(1 for p in p2s if p not in hok_p2)
        new_p3 = sum(1 for s in syms if s not in hok_p3)
        return rarity*2 + new_p1*3 + new_p2*2 + new_p3*1

    for d in non_hok:
        d["score"] = score_dam(d)
        d["new_p1"]= sum(1 for r in d["recs"] if sym_parts(r["symbol"])[0] not in hok_p1)
        d["new_p3"]= sum(1 for r in d["recs"] if r["symbol"] not in hok_p3)

    # 上位100を都道府県分散を考慮して選定
    # まず純スコア降順でソート、都道府県ごとに最大5件に制限
    candidates = sorted(non_hok, key=lambda d:-d["score"])
    selected = []; pref_cnt=Counter()
    for d in candidates:
        if len(selected)>=100: break
        if pref_cnt[d["pref"]]<5:
            selected.append(d); pref_cnt[d["pref"]]+=1

    # もし100件未満なら制限解除で補充
    if len(selected)<100:
        added={id(d) for d in selected}
        for d in candidates:
            if len(selected)>=100: break
            if id(d) not in added:
                selected.append(d); added.add(id(d))

    row=2
    # サマリー
    sel_syms=[r["symbol"] for d in selected for r in d["recs"]]
    sel_p1=set(sym_parts(s)[0] for s in sel_syms)
    sel_p2=set(sym_parts(s)[1] for s in sel_syms)
    sel_p3=set(sel_syms)
    b_all=[b_score(r["bearing_cap"]) for d in selected for r in d["recs"] if b_score(r["bearing_cap"])]
    p_all=[p_score(r["permeability"]) for d in selected for r in d["recs"] if p_score(r["permeability"])]

    summary=[
        ("選定ダム数",len(selected)),
        ("ユニークSymbol（3項目）数",len(sel_p3)),
        ("ユニーク1項目数",len(sel_p1)),
        ("ユニーク2項目数",len(sel_p2)),
        ("平均強度スコア",avg(b_all)),
        ("平均透水性スコア",avg(p_all)),
        ("北海道未カバーSymbol含有",sum(1 for s in sel_p3 if s not in hok_p3)),
        ("都道府県数",len(pref_cnt)),
    ]
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=14
    for label,val in summary:
        ws.cell(row,1,label).font=BOLD_FONT; ws.cell(row,1).alignment=LEFT
        ws.cell(row,2,val).font=BODY_FONT;  ws.cell(row,2).alignment=RIGHT
        row+=1
    row+=1

    # 選定ダム一覧
    sec_title(ws,row,1,16,"▼ 選定100ダム一覧（スコア降順）")
    row+=1
    hdrs=["順位","ダム名","都道府県","堤高(m)","管理者コード",
          "Symbol-1","1項目-1","Symbol-2","Symbol-3",
          "選定スコア","新規1項目数","新規Symbol数",
          "平均強度","平均透水性","平均リスク","主リスク"]
    widths=[6,16,12,8,10,18,12,16,16,10,10,10,10,10,10,22]
    for ci,(h,w) in enumerate(zip(hdrs,widths),1): hdr(ws,row,ci,h,w)
    row+=1

    for rank,d in enumerate(selected,1):
        syms=[r["symbol"] for r in d["recs"]]
        bs=[b_score(r["bearing_cap"]) for r in d["recs"] if b_score(r["bearing_cap"])]
        ps=[p_score(r["permeability"]) for r in d["recs"] if p_score(r["permeability"])]
        risks=[risk_rank(r["bearing_cap"],r["permeability"]) for r in d["recs"]]
        risk_top=Counter(r["main_risk"] for r in d["recs"]).most_common(1)
        fill=RED_FILL if d["new_p1"]>0 else ALT_FILL if row%2==0 else None
        s1=syms[0] if syms else ""
        vals=[rank,d["name"],d["pref"],d["height"],d["mgr_code"],
              s1,sym_parts(s1)[0] if s1 else "",
              syms[1] if len(syms)>1 else "",
              syms[2] if len(syms)>2 else "",
              round(d["score"],3),d["new_p1"],d["new_p3"],
              avg(bs),avg(ps),avg(risks),
              risk_top[0][0] if risk_top else ""]
        for ci,v in enumerate(vals,1):
            al=RIGHT if ci in(1,4,5,10,11,12,13,14,15) else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row+=1

    ws.freeze_panes="A3"
    ws.auto_filter.ref=f"A2:{scol(len(hdrs))}2"


# ─── S7: カバレッジ比較 ─────────────────────────────────────
def write_s7(wb, dams, glossary):
    ws = wb.create_sheet("S7_カバレッジ比較")
    ws.sheet_view.showGridLines = False
    sheet_title(ws,1,1,14,"■ カバレッジ比較：北海道60 vs 全国100（Symbol・強度透水性ギャップ）")

    # 北海道60
    hok=[d for d in dams if d["pref"].startswith("北海道") and d["mgr_code"]==1 and d["recs"]]
    hok60=sorted(hok,key=lambda d:-(d["height"] or 0))[:60]
    hok_p1=set(sym_parts(r["symbol"])[0] for d in hok60 for r in d["recs"])
    hok_p2=set(sym_parts(r["symbol"])[1] for d in hok60 for r in d["recs"])
    hok_p3=set(r["symbol"]              for d in hok60 for r in d["recs"])

    # 全国100（北海道除く）: S6と同じロジック再計算
    non_hok=[d for d in dams if not d["pref"].startswith("北海道") and d["recs"]]
    all_sym_cnt=Counter(r["symbol"] for d in dams for r in d["recs"])
    def score_dam(d):
        syms=[r["symbol"] for r in d["recs"]]
        p1s=[sym_parts(s)[0] for s in syms]; p2s=[sym_parts(s)[1] for s in syms]
        return (sum(1.0/all_sym_cnt[s] for s in syms)*2 +
                sum(1 for p in p1s if p not in hok_p1)*3 +
                sum(1 for p in p2s if p not in hok_p2)*2 +
                sum(1 for s in syms if s not in hok_p3)*1)
    candidates=sorted(non_hok,key=lambda d:-score_dam(d))
    sel=[]; pc=Counter()
    for d in candidates:
        if len(sel)>=100: break
        if pc[d["pref"]]<5: sel.append(d); pc[d["pref"]]+=1
    if len(sel)<100:
        added={id(d) for d in sel}
        for d in candidates:
            if len(sel)>=100: break
            if id(d) not in added: sel.append(d); added.add(id(d))

    nat_p1=set(sym_parts(r["symbol"])[0] for d in sel for r in d["recs"])
    nat_p2=set(sym_parts(r["symbol"])[1] for d in sel for r in d["recs"])
    nat_p3=set(r["symbol"]              for d in sel for r in d["recs"])
    all_p1=set(sym_parts(r["symbol"])[0] for d in dams for r in d["recs"])
    all_p2=set(sym_parts(r["symbol"])[1] for d in dams for r in d["recs"])
    all_p3=set(r["symbol"]              for d in dams for r in d["recs"])
    combined_p1=hok_p1|nat_p1; combined_p3=hok_p3|nat_p3

    row=2
    # カバレッジサマリー表
    sec_title(ws,row,1,6,"▼ カバレッジサマリー")
    row+=1
    for ci,t in enumerate(["","北海道60","全国100","北海道60+全国100","全国全体","カバー率(%)"],1):
        sub(ws,row,ci,t,20 if ci==1 else 14)
    row+=1
    metrics=[
        ("1項目(地質時代)種類数",len(hok_p1),len(nat_p1),len(combined_p1),len(all_p1)),
        ("2項目種類数",          len(hok_p2),len(nat_p2),len(hok_p2|nat_p2),len(all_p2)),
        ("3項目(フルSymbol)種類数",len(hok_p3),len(nat_p3),len(combined_p3),len(all_p3)),
    ]
    for label,h,n,c,total in metrics:
        fill=ALT_FILL if row%2==0 else None
        for ci,v in enumerate([label,h,n,c,total,pct(c,total)],1):
            al=RIGHT if ci>1 else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row+=1

    row+=1
    # 北海道60で未カバー・全国100でカバーするSymbol（P3）
    gap_covered=nat_p3-hok_p3
    gap_still=(all_p3-hok_p3)-nat_p3

    sec_title(ws,row,1,10,"▼ 全国100で補完するSymbol（北海道60未カバー → 全国100でカバー）")
    row+=1
    for ci,t in enumerate(["Symbol","1項目","2項目","geo_era","geo_rock",
                            "全国層数","bearing_cap","permeability","リスクスコア","main_risk"],1):
        sub(ws,row,ci,t,18 if ci in(1,3) else 10)
    row+=1
    for sym in sorted(gap_covered, key=lambda s:-all_sym_cnt.get(s,0))[:80]:
        rec=next((r for r in glossary.values() if r["symbol"]==sym),{})
        rr=risk_rank(rec.get("bearing_cap",""),rec.get("permeability",""))
        fill=RED_FILL if rr>=6 else YEL_FILL if rr>=4 else ALT_FILL if row%2==0 else None
        vals=[sym,sym_parts(sym)[0],sym_parts(sym)[1],
              rec.get("geo_era",""),rec.get("geo_rock",""),
              all_sym_cnt.get(sym,0),
              rec.get("bearing_cap",""),rec.get("permeability",""),rr,rec.get("main_risk","")]
        for ci,v in enumerate(vals,1):
            al=RIGHT if ci in(6,9) else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row+=1

    row+=1
    # 160ダム合計でも未カバーのSymbol
    sec_title(ws,row,1,10,f"▼ 北海道60+全国100（計160ダム）でも未カバーのSymbol（{len(gap_still)}種）")
    row+=1
    for ci,t in enumerate(["Symbol","1項目","geo_era","geo_rock",
                            "全国層数","bearing_cap","permeability","リスクスコア","main_risk"],1):
        sub(ws,row,ci,t,18 if ci==1 else 10)
    row+=1
    for sym in sorted(gap_still,key=lambda s:-all_sym_cnt.get(s,0)):
        rec=next((r for r in glossary.values() if r["symbol"]==sym),{})
        rr=risk_rank(rec.get("bearing_cap",""),rec.get("permeability",""))
        fill=RED_FILL if rr>=6 else ALT_FILL if row%2==0 else None
        vals=[sym,sym_parts(sym)[0],rec.get("geo_era",""),rec.get("geo_rock",""),
              all_sym_cnt.get(sym,0),rec.get("bearing_cap",""),
              rec.get("permeability",""),rr,rec.get("main_risk","")]
        for ci,v in enumerate(vals,1):
            al=RIGHT if ci in(5,8) else LEFT
            body(ws,row,ci,v,align=al,fill=fill)
        row+=1

    ws.freeze_panes="A3"


# ─── メイン ─────────────────────────────────────────────────
def parse_args():
    p=argparse.ArgumentParser(description="全国ダム地質DB 統計分析 v2")
    p.add_argument("--input",  required=True)
    p.add_argument("--output", required=True)
    return p.parse_args()

def main():
    args=parse_args()
    print(f"読み込み: {args.input}")
    wb=load_workbook(args.input)
    dams,glossary=load_data(wb)
    has=[d for d in dams if d["recs"]]
    print(f"総ダム数:{len(dams)} / データあり:{len(has)} / Symbol種:{len(set(r['symbol'] for d in has for r in d['recs']))}")

    print("S1 Symbol階層分析...")
    write_s1(wb,dams,glossary)
    print("S2 強度透水性マトリクス...")
    write_s2(wb,dams,glossary)
    print("S3 Symbol類似グループ...")
    write_s3(wb,dams,glossary)
    print("S4 2項目組合せ...")
    write_s4(wb,dams,glossary)
    print("S5 北海道60ダム...")
    write_s5(wb,dams,glossary)
    print("S6 全国100ダム選定...")
    write_s6(wb,dams,glossary)
    print("S7 カバレッジ比較...")
    write_s7(wb,dams,glossary)

    wb.save(args.output)
    print(f"\n保存完了: {args.output}")

if __name__=="__main__":
    main()
